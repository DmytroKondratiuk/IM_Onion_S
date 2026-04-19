"""
Exports full IM Onion-S simulation results to xlsx (uk / en).

All openpyxl charts have `.title = None` — the title is carried by the sheet
header and by the caption in the consuming document (md / docx). This keeps
figures clean for publication use.
"""

from __future__ import annotations
import os
from datetime import datetime
from typing import Dict, Sequence, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, BarChart, RadarChart, Reference
from openpyxl.utils import get_column_letter

from i18n import get_labels, scenario_label


FONT_TITLE  = Font(name="Arial", size=12, bold=True)
FONT_HEAD   = Font(name="Arial", size=10, bold=True)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_SOURCE = Font(name="Arial", size=9, italic=True, color="555555")

FILL_HEADER = PatternFill("solid", fgColor="D9E5F5")
FILL_TITLE  = PatternFill("solid", fgColor="F5F5DC")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COMPONENT_LABELS = {
    "C":  "DM_C",  "L1": "DM_L₁", "L2": "DM_L₂",
    "L3": "DM_L₃", "L4": "DM_L₄", "I":  "DM_I",  "S":  "DM_S",
}


def _write_title(ws, title: str, row: int = 1, col: int = 1,
                 span_cols: int = 6) -> None:
    ws.cell(row=row, column=col, value=title).font = FONT_TITLE
    ws.cell(row=row, column=col).fill = FILL_TITLE
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span_cols - 1)


def _write_source(ws, source: str, row: int, col: int = 1,
                  span_cols: int = 6) -> None:
    ws.cell(row=row, column=col, value=source).font = FONT_SOURCE
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span_cols - 1)


def _write_headers(ws, headers, row: int, col: int = 1) -> None:
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col + i, value=h)
        cell.font = FONT_HEAD
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER


def _autosize_columns(ws, min_width: int = 10, max_width: int = 28) -> None:
    for col_cells in ws.columns:
        try:
            letter = col_cells[0].column_letter
        except AttributeError:
            continue
        max_len = 0
        for c in col_cells:
            if c.value is None:
                continue
            max_len = max(max_len, min(len(str(c.value)), max_width))
        ws.column_dimensions[letter].width = max(min_width, max_len + 2)


# ───────────────────── individual sheets ─────────────────────

def _sheet_metadata(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_metadata"])
    _write_title(ws, L["title_metadata"], 1, 1, 4)
    rows = [
        (L["meta_run_date"],      source_label),
        (L["meta_final_alpha"],   f"{report['calibration']['final_alpha']:.4f}"),
        (L["meta_final_delta"],   f"{report['calibration']['final_delta']:.4f}"),
        (L["meta_mean_alpha"],    f"{report['calibration']['mean_alpha']:.4f} ± "
                                  f"{report['calibration']['std_alpha']:.4f}"),
        (L["meta_mean_delta"],    f"{report['calibration']['mean_delta']:.4f} ± "
                                  f"{report['calibration']['std_delta']:.4f}"),
        (L["meta_n_enterprises"], report["config"]["n_enterprises"]),
        (L["meta_n_runs"],        report["config"]["n_runs"]),
        (L["meta_horizon"],       report["config"]["n_years"]),
        (L["meta_step"],          report["config"]["dt"]),
        (L["meta_exec_time"],     report.get("execution_time_seconds", L["n_a"])),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = FONT_HEAD
        ws.cell(row=r, column=2, value=value).font = FONT_NORMAL
        r += 1
    _autosize_columns(ws)


def _sheet_calibration(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_calibration"])
    _write_title(ws, L["title_calibration"], 1, 1, 8)
    _write_headers(ws, [L["col_run"], L["col_seed"], "α*", "δ*",
                        "J(α*,δ*)", "C₁", "C₂", "C₃"], 3)
    runs = report["calibration"]["individual_runs"]
    r = 4
    for run in runs:
        ws.cell(row=r, column=1, value=run["run_id"] + 1)
        ws.cell(row=r, column=2, value=42 + run["run_id"] * 137)
        ws.cell(row=r, column=3, value=float(run["best_alpha"]))
        ws.cell(row=r, column=4, value=float(run["best_delta"]))
        ws.cell(row=r, column=5, value=float(run["best_score"]))
        ws.cell(row=r, column=6, value=float(run["c1_correlation"]))
        ws.cell(row=r, column=7, value=float(run["c2_entropy"]))
        ws.cell(row=r, column=8, value=float(run["c3_misclass"]))
        r += 1
    ws.cell(row=r, column=1, value=L["col_mean"]).font = FONT_HEAD
    ws.cell(row=r, column=3, value=report["calibration"]["mean_alpha"])
    ws.cell(row=r, column=4, value=report["calibration"]["mean_delta"])
    r += 1
    ws.cell(row=r, column=1, value=L["col_stdev"]).font = FONT_HEAD
    ws.cell(row=r, column=3, value=report["calibration"]["std_alpha"])
    ws.cell(row=r, column=4, value=report["calibration"]["std_delta"])
    r += 2
    _write_source(ws, L["source_simulation"].format(ts=source_label), r, 1, 8)
    for row in range(4, 4 + len(runs)):
        for col in range(3, 9):
            ws.cell(row=row, column=col).number_format = "0.0000"
    _autosize_columns(ws)


def _sheet_calib_heatmap(wb, calib_grid, alpha_range, delta_range,
                         source_label, L):
    ws = wb.create_sheet(L["sheet_calib_heatmap"])
    _write_title(ws, L["title_calib_heatmap"], 1, 1, len(delta_range) + 1)
    ws.cell(row=3, column=1, value=L["col_alpha_delta"]).font = FONT_HEAD
    ws.cell(row=3, column=1).fill = FILL_HEADER
    for j, d in enumerate(delta_range):
        c = ws.cell(row=3, column=2 + j, value=float(d))
        c.font = FONT_HEAD; c.fill = FILL_HEADER
        c.number_format = "0.000"; c.alignment = ALIGN_CENTER
    r = 4
    for i, a in enumerate(alpha_range):
        ws.cell(row=r, column=1, value=float(a)).font = FONT_HEAD
        ws.cell(row=r, column=1).number_format = "0.00"
        for j in range(len(delta_range)):
            cell = ws.cell(row=r, column=2 + j, value=float(calib_grid[i, j]))
            cell.number_format = "0.0000"
        r += 1
    last_col_letter = get_column_letter(1 + len(delta_range))
    rng = f"B4:{last_col_letter}{4 + len(alpha_range) - 1}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))
    _write_source(ws, L["source_simulation"].format(ts=source_label),
                  4 + len(alpha_range) + 1, 1, len(delta_range) + 1)
    _autosize_columns(ws)


def _sheet_degradation(wb, source_label, L):
    from onion_s_simulation import CONFIG
    ws = wb.create_sheet(L["sheet_degradation"])
    _write_title(ws, L["title_degradation"], 1, 1, 3)
    _write_headers(ws, [L["col_component"], L["col_dk"], L["col_justification"]], 3)
    r = 4
    for k in ["C", "L1", "L2", "L3", "L4", "I", "S"]:
        ws.cell(row=r, column=1, value=L[f"comp_name_{k}"])
        ws.cell(row=r, column=2, value=CONFIG["depreciation_by_layer"][k]
                ).number_format = "0.000"
        ws.cell(row=r, column=3, value=L[f"reason_{k}"])
        r += 1
    r += 1
    _write_source(ws, L["source_developed"].format(ts=source_label), r, 1, 3)
    _autosize_columns(ws)


def _sheet_scenarios_admi(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_scenarios_admi"])
    _write_title(ws, L["title_scenarios_admi"], 1, 1, 7)
    scenarios = list(report["scenarios"].keys())
    n_years = len(report["scenarios"][scenarios[0]]["admi_mean"])
    headers = [L["ax_year"]] + [scenario_label(s, L) for s in scenarios]
    _write_headers(ws, headers, 3)
    for year in range(n_years):
        ws.cell(row=4 + year, column=1, value=year)
        for i, s in enumerate(scenarios):
            ws.cell(row=4 + year, column=2 + i,
                    value=float(report["scenarios"][s]["admi_mean"][year])
                    ).number_format = "0.0000"
    next_row = 4 + n_years + 1
    ws.cell(row=next_row, column=1, value="StDev").font = FONT_HEAD
    next_row += 1
    _write_headers(ws, headers, next_row)
    for year in range(n_years):
        ws.cell(row=next_row + 1 + year, column=1, value=year)
        for i, s in enumerate(scenarios):
            ws.cell(row=next_row + 1 + year, column=2 + i,
                    value=float(report["scenarios"][s]["admi_std"][year])
                    ).number_format = "0.0000"
    next_row += n_years + 2
    _write_source(ws, L["source_simulation"].format(ts=source_label),
                  next_row, 1, len(headers))
    _autosize_columns(ws)


def _sheet_scenarios_dm(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_scenarios_dm"])
    _write_title(ws, L["title_scenarios_dm"], 1, 1, 9)
    scenarios = list(report["scenarios"].keys())
    components = ["C", "L1", "L2", "L3", "L4", "I", "S"]
    n_years = len(report["scenarios"][scenarios[0]]["admi_mean"])
    r = 3
    for s in scenarios:
        ws.cell(row=r, column=1, value=scenario_label(s, L)).font = FONT_HEAD
        ws.cell(row=r, column=1).fill = FILL_TITLE
        r += 1
        headers = [L["ax_year"]] + [COMPONENT_LABELS[c] for c in components]
        _write_headers(ws, headers, r)
        r += 1
        for year in range(n_years):
            ws.cell(row=r, column=1, value=year)
            dm_comp = report["scenarios"][s].get("dm_components_mean", {})
            for i, c in enumerate(components):
                vals = dm_comp.get(c, [None] * n_years)
                val = vals[year] if vals[year] is not None else None
                ws.cell(row=r, column=2 + i,
                        value=float(val) if val is not None else None
                        ).number_format = "0.0000"
            r += 1
        r += 1
    _write_source(ws, L["source_simulation"].format(ts=source_label), r, 1, 8)
    _autosize_columns(ws)


def _sheet_level_distributions(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_level_dist"])
    _write_title(ws, L["title_level_dist"], 1, 1, 7)
    scenarios = list(report["scenarios"].keys())
    n_years = len(report["scenarios"][scenarios[0]]["level_distributions"])
    r = 3
    for s in scenarios:
        ws.cell(row=r, column=1, value=scenario_label(s, L)).font = FONT_HEAD
        ws.cell(row=r, column=1).fill = FILL_TITLE
        r += 1
        _write_headers(ws, [L["ax_year"],
                            f"{L['level_prefix']} 1", f"{L['level_prefix']} 2",
                            f"{L['level_prefix']} 3", f"{L['level_prefix']} 4",
                            f"{L['level_prefix']} 5"], r)
        r += 1
        for year in range(n_years):
            ws.cell(row=r, column=1, value=year)
            dist = report["scenarios"][s]["level_distributions"][year]
            for i, v in enumerate(dist):
                ws.cell(row=r, column=2 + i, value=float(v)
                        ).number_format = "0.00%"
            r += 1
        r += 1
    _write_source(ws, L["source_simulation"].format(ts=source_label), r, 1, 6)
    _autosize_columns(ws)


def _sheet_markov_matrix(wb, scenario_name, P, source_label, L):
    sheet_name = L.get(f"sheet_markov_{scenario_name}",
                        f"Markov {scenario_name}")
    ws = wb.create_sheet(sheet_name)
    _write_title(ws, L["title_markov"].format(scenario=scenario_label(scenario_name, L)),
                 1, 1, 6)
    _write_headers(ws, [L["col_from_to_level"], "1", "2", "3", "4", "5"], 3)
    for i in range(5):
        ws.cell(row=4 + i, column=1, value=L[f"level_{i+1}_full"]).font = FONT_HEAD
        for j in range(5):
            c = ws.cell(row=4 + i, column=2 + j, value=float(P[i, j]))
            c.number_format = "0.000"
            c.alignment = ALIGN_CENTER
    ws.conditional_formatting.add("B4:F8", ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        mid_type="num", mid_value=0.5, mid_color="9DC3E6",
        end_type="num", end_value=1, end_color="1F4E78"))
    _write_source(ws, L["source_simulation"].format(ts=source_label), 10, 1, 6)
    _autosize_columns(ws)


def _sheet_stationary(wb, cv_results, source_label, L):
    ws = wb.create_sheet(L["sheet_stationary"])
    _write_title(ws, L["title_stationary"], 1, 1, 6)
    _write_headers(ws, [L["ax_scenario"], "π₁", "π₂", "π₃", "π₄", "π₅"], 3)
    r = 4
    for s, cv in cv_results.items():
        ws.cell(row=r, column=1, value=scenario_label(s, L))
        for i, v in enumerate(cv["stationary_distribution"]):
            ws.cell(row=r, column=2 + i, value=float(v)).number_format = "0.000"
        r += 1
    r += 1
    _write_source(ws, L["source_simulation"].format(ts=source_label), r, 1, 6)
    _autosize_columns(ws)


def _sheet_hitting_times(wb, cv_results, source_label, L):
    ws = wb.create_sheet(L["sheet_hitting_times"])
    _write_title(ws, L["title_hitting_times"], 1, 1, len(cv_results) + 1)
    all_transitions = set()
    for cv in cv_results.values():
        all_transitions.update(cv.get("hitting_times", {}).keys())
    all_transitions = sorted(all_transitions)
    headers = [L["col_transition"]] + [scenario_label(s, L) for s in cv_results.keys()]
    _write_headers(ws, headers, 3)
    r = 4
    for tr in all_transitions:
        ws.cell(row=r, column=1, value=tr)
        for i, (s, cv) in enumerate(cv_results.items()):
            t = cv.get("hitting_times", {}).get(tr)
            if t is None or t > 100:
                ws.cell(row=r, column=2 + i, value=L["dash"]).alignment = ALIGN_CENTER
            else:
                ws.cell(row=r, column=2 + i, value=float(t)).number_format = "0.0"
        r += 1
    r += 1
    _write_source(ws, L["source_hitting"].format(ts=source_label),
                  r, 1, len(headers))
    _autosize_columns(ws)


def _sheet_multihorizon(wb, mh_results, source_label, L):
    if not mh_results:
        return
    ws = wb.create_sheet(L["sheet_multihorizon"])
    _write_title(ws, L["title_multihorizon"], 1, 1, 6)
    _write_headers(ws, [L["ax_scenario"], L["col_delta_admi_5"],
                        L["col_delta_admi_10"], L["col_delta_admi_15"],
                        L["col_ratio_15_5"]], 3)
    scenarios = list(mh_results["scenarios_by_horizon"][5].keys())
    r = 4
    for s in scenarios:
        ws.cell(row=r, column=1, value=scenario_label(s, L))
        admi_changes = []
        for h in (5, 10, 15):
            data = mh_results["scenarios_by_horizon"][h][s]
            ch = float(data["admi_mean"][-1] - data["admi_mean"][0])
            admi_changes.append(ch)
        for i, ch in enumerate(admi_changes):
            ws.cell(row=r, column=2 + i, value=ch).number_format = "0.000"
        if admi_changes[0] != 0:
            ratio = admi_changes[2] / admi_changes[0]
            ws.cell(row=r, column=5,
                    value=f"{abs(ratio):.1f}×" if ratio != 0 else L["n_a"])
        r += 1
    r += 1
    _write_source(ws, L["source_multihorizon"].format(ts=source_label), r, 1, 5)
    _autosize_columns(ws)


def _sheet_crossvalidation(wb, cv_results, source_label, L):
    ws = wb.create_sheet(L["sheet_crossval"])
    _write_title(ws, L["title_crossval"], 1, 1, 5)
    r = 3
    for scenario_name, cv in cv_results.items():
        ws.cell(row=r, column=1, value=scenario_label(scenario_name, L)).font = FONT_HEAD
        ws.cell(row=r, column=1).fill = FILL_TITLE
        r += 1
        _write_headers(ws, [L["ax_level"], L["col_sd_empirical"],
                            L["col_markov_predicted"], L["col_difference_pp"]], r)
        r += 1
        sd = cv["empirical_distribution_year5"]
        mk = cv["markov_predicted_year5"]
        for i in range(5):
            ws.cell(row=r, column=1, value=f"{L['level_prefix']} {i+1}")
            ws.cell(row=r, column=2, value=float(sd[i])).number_format = "0.00%"
            ws.cell(row=r, column=3, value=float(mk[i])).number_format = "0.00%"
            diff = (sd[i] - mk[i]) * 100
            ws.cell(row=r, column=4, value=round(diff, 1)).number_format = "0.0"
            r += 1
        ws.cell(row=r, column=1, value=L["col_l1_deviation"]).font = FONT_HEAD
        ws.cell(row=r, column=4, value=float(cv["l1_deviation_predicted"])
                ).number_format = "0.000"
        r += 2
    _write_source(ws, L["source_simulation"].format(ts=source_label), r, 1, 4)
    _autosize_columns(ws)


def _sheet_sensitivity(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_sensitivity"])
    _write_title(ws, L["title_sensitivity"], 1, 1, 5)
    _write_headers(ws, [L["ax_parameter"], L["col_base_value"],
                        L["col_metric_low"], L["col_metric_high"],
                        L["col_sensitivity"]], 3)
    sens = report["sensitivity"]
    sorted_params = sorted(sens.items(), key=lambda x: x[1]["sensitivity"], reverse=True)
    r = 4
    for param, data in sorted_params:
        ws.cell(row=r, column=1, value=param)
        ws.cell(row=r, column=2, value=float(data["base_value"])).number_format = "0.000"
        ws.cell(row=r, column=3, value=float(data["metric_at_low"])).number_format = "0.000"
        ws.cell(row=r, column=4, value=float(data["metric_at_high"])).number_format = "0.000"
        ws.cell(row=r, column=5, value=float(data["sensitivity"])).number_format = "0.0000"
        r += 1
    r += 1
    _write_source(ws, L["source_simulation"].format(ts=source_label), r, 1, 5)
    _autosize_columns(ws)


# ───────────────────── chart sheets (charts without titles) ─────────────────────

def _chart_trajectories(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_fig_trajectories"])
    _write_title(ws, L["title_fig_trajectories"], 1, 1, 6)
    scenarios = list(report["scenarios"].keys())
    n_years = len(report["scenarios"][scenarios[0]]["admi_mean"])
    headers = [L["ax_year"]] + [scenario_label(s, L) for s in scenarios]
    _write_headers(ws, headers, 3)
    for y in range(n_years):
        ws.cell(row=4 + y, column=1, value=y)
        for i, s in enumerate(scenarios):
            ws.cell(row=4 + y, column=2 + i,
                    value=float(report["scenarios"][s]["admi_mean"][y])
                    ).number_format = "0.000"
    ch = LineChart()
    ch.title = None  # no title (caption provided in publication)
    ch.y_axis.title = L["ax_admi"]
    ch.x_axis.title = L["ax_year"]
    data = Reference(ws, min_col=2, min_row=3,
                     max_col=1 + len(scenarios), max_row=3 + n_years)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n_years)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height = 11; ch.width = 18
    ws.add_chart(ch, "H3")
    _write_source(ws, L["source_simulation"].format(ts=source_label),
                  5 + n_years, 1, 6)
    _autosize_columns(ws)


def _chart_profiles(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_fig_profiles"])
    _write_title(ws, L["title_fig_profiles"], 1, 1, 7)
    scenarios = list(report["scenarios"].keys())
    components = ["C", "L1", "L2", "L3", "L4", "I", "S"]
    headers = [L["col_component"]] + [scenario_label(s, L) for s in scenarios]
    _write_headers(ws, headers, 3)
    for i, c in enumerate(components):
        ws.cell(row=4 + i, column=1, value=COMPONENT_LABELS[c])
        for j, s in enumerate(scenarios):
            dm_comp = report["scenarios"][s].get("dm_components_mean", {})
            vals = dm_comp.get(c, [])
            v = vals[-1] if vals else None
            ws.cell(row=4 + i, column=2 + j,
                    value=float(v) if v is not None else None
                    ).number_format = "0.000"
    ch = RadarChart()
    ch.type = "filled"; ch.style = 26
    ch.title = None  # no title (caption provided in publication)
    data = Reference(ws, min_col=2, min_row=3,
                     max_col=1 + len(scenarios), max_row=3 + len(components))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(components))
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height = 12; ch.width = 14
    ws.add_chart(ch, "J3")
    _write_source(ws, L["source_simulation"].format(ts=source_label),
                  5 + len(components), 1, 7)
    _autosize_columns(ws)


def _chart_multihorizon(wb, mh_results, source_label, L):
    if not mh_results:
        return
    ws = wb.create_sheet(L["sheet_fig_multihorizon"])
    _write_title(ws, L["title_fig_multihorizon"], 1, 1, 5)
    scenarios = list(mh_results["scenarios_by_horizon"][5].keys())
    headers = [L["ax_scenario"], L["col_horizon_5"],
               L["col_horizon_10"], L["col_horizon_15"]]
    _write_headers(ws, headers, 3)
    for i, s in enumerate(scenarios):
        ws.cell(row=4 + i, column=1, value=scenario_label(s, L))
        for j, h in enumerate((5, 10, 15)):
            data = mh_results["scenarios_by_horizon"][h][s]
            ch_val = float(data["admi_mean"][-1] - data["admi_mean"][0])
            ws.cell(row=4 + i, column=2 + j, value=ch_val).number_format = "0.000"
    ch = BarChart()
    ch.type = "col"; ch.style = 11
    ch.title = None  # no title (caption provided in publication)
    ch.y_axis.title = L["ax_delta_admi"]
    ch.x_axis.title = L["ax_scenario"]
    data = Reference(ws, min_col=2, min_row=3, max_col=4,
                     max_row=3 + len(scenarios))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(scenarios))
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height = 11; ch.width = 16
    ws.add_chart(ch, "G3")
    _write_source(ws, L["source_multihorizon"].format(ts=source_label),
                  5 + len(scenarios), 1, 4)
    _autosize_columns(ws)


def _chart_crossvalidation(wb, cv_results, source_label, L):
    ws = wb.create_sheet(L["sheet_fig_crossval"])
    _write_title(ws, L["title_fig_crossval"], 1, 1, 4)
    cv = cv_results.get("uniform") or list(cv_results.values())[0]
    headers = [L["ax_level"], L["col_sd_empirical"], L["col_markov_predicted"]]
    _write_headers(ws, headers, 3)
    for i in range(5):
        ws.cell(row=4 + i, column=1, value=f"{L['level_prefix']} {i+1}")
        ws.cell(row=4 + i, column=2,
                value=float(cv["empirical_distribution_year5"][i])
                ).number_format = "0.000"
        ws.cell(row=4 + i, column=3,
                value=float(cv["markov_predicted_year5"][i])
                ).number_format = "0.000"
    ch = BarChart()
    ch.type = "col"; ch.style = 12
    ch.title = None  # no title (caption provided in publication)
    ch.y_axis.title = L["ax_enterprise_share"]
    ch.x_axis.title = L["ax_level"]
    data = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=8)
    cats = Reference(ws, min_col=1, min_row=4, max_row=8)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height = 10; ch.width = 15
    ws.add_chart(ch, "F3")
    _write_source(ws, L["source_crossval"].format(ts=source_label,
                                                    l1=cv['l1_deviation_predicted']),
                  10, 1, 4)
    _autosize_columns(ws)


def _chart_sensitivity(wb, report, source_label, L):
    ws = wb.create_sheet(L["sheet_fig_sensitivity"])
    _write_title(ws, L["title_fig_sensitivity"], 1, 1, 3)
    headers = [L["ax_parameter"], L["col_sensitivity"]]
    _write_headers(ws, headers, 3)
    sens = report["sensitivity"]
    sorted_params = sorted(sens.items(), key=lambda x: x[1]["sensitivity"], reverse=True)
    for i, (param, data) in enumerate(sorted_params):
        ws.cell(row=4 + i, column=1, value=param)
        ws.cell(row=4 + i, column=2, value=float(data["sensitivity"])
                ).number_format = "0.0000"
    ch = BarChart()
    ch.type = "bar"; ch.style = 10
    ch.title = None  # no title (caption provided in publication)
    ch.y_axis.title = L["ax_parameter"]
    ch.x_axis.title = L["col_sensitivity"]
    data = Reference(ws, min_col=2, min_row=3, max_col=2,
                     max_row=3 + len(sorted_params))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(sorted_params))
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height = 11; ch.width = 16
    ws.add_chart(ch, "E3")
    _write_source(ws, L["source_simulation"].format(ts=source_label),
                  5 + len(sorted_params), 1, 3)
    _autosize_columns(ws)


# ───────────────────── main entry point ─────────────────────

def export_to_xlsx(report, cv_results, multi_horizon, output_path,
                    calib_grid=None, alpha_range=None, delta_range=None,
                    timestamp=None, lang: str = "uk") -> None:
    """Exports the full simulation report to xlsx."""
    L = get_labels(lang)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    if timestamp is None:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    source_label = timestamp

    _sheet_metadata(wb, report, source_label, L)
    _sheet_calibration(wb, report, source_label, L)
    if calib_grid is not None and alpha_range is not None and delta_range is not None:
        _sheet_calib_heatmap(wb, calib_grid, alpha_range, delta_range, source_label, L)
    _sheet_degradation(wb, source_label, L)
    _sheet_scenarios_admi(wb, report, source_label, L)
    _sheet_scenarios_dm(wb, report, source_label, L)
    _sheet_level_distributions(wb, report, source_label, L)
    for scenario_name, cv in cv_results.items():
        P = np.array(cv["transition_matrix"])
        _sheet_markov_matrix(wb, scenario_name, P, source_label, L)
    _sheet_stationary(wb, cv_results, source_label, L)
    _sheet_hitting_times(wb, cv_results, source_label, L)
    _sheet_multihorizon(wb, multi_horizon, source_label, L)
    _sheet_crossvalidation(wb, cv_results, source_label, L)
    _sheet_sensitivity(wb, report, source_label, L)

    _chart_trajectories(wb, report, source_label, L)
    _chart_profiles(wb, report, source_label, L)
    _chart_multihorizon(wb, multi_horizon, source_label, L)
    _chart_crossvalidation(wb, cv_results, source_label, L)
    _chart_sensitivity(wb, report, source_label, L)

    wb.save(output_path)
